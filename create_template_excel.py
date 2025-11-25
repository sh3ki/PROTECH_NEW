"""
Script to create a properly formatted Excel template for user import
Run this once to generate the template file with blue header styling
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Users Import Template"

# Define styles (matching export format)
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center")

# Set column widths (matching export)
column_widths = [8, 15, 25, 15, 15, 15, 15, 12, 15, 18]
for idx, width in enumerate(column_widths, start=1):
    ws.column_dimensions[chr(64 + idx)].width = width

# Headers
headers = ["ID", "Username", "Email", "First Name", "Last Name", "Middle Name", "Role", "Status", "Password", "Created Date"]

# Add headers with styling
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_alignment
    cell.border = border_style

# Set header row height
ws.row_dimensions[1].height = 25

# Sample data rows
sample_data = [
    [146, "john.teacher", "john.teacher@protech.com", "John", "Teacher", "M", "Teacher", "Active", "Password123", "2025-11-21 16:41:29"],
    [145, "jane.admin", "jane.admin@protech.com", "Jane", "Admin", "S", "Administrator", "Active", "AdminPass123", "2025-11-21 16:00:37"],
    [144, "robert.principal", "robert.principal@protech.com", "Robert", "Principal", "K", "Principal", "Active", "PrincipalPass123", "2025-11-21 15:48:27"],
    [140, "mary.registrar", "mary.registrar@protech.com", "Mary", "Registrar", "L", "Registrar", "Active", "RegistrarPass123", "2025-11-21 14:55:42"],
]

# Add data rows with styling
for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.border = border_style
        
        # Apply alignment
        if col_idx in [1, 8]:  # ID and Status columns - center
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment

# Freeze header row
ws.freeze_panes = ws['A2']

# Save the file
wb.save('static/templates/users_import_template.xlsx')
print("✅ Excel template created successfully: static/templates/users_import_template.xlsx")
print("📋 Template includes:")
print("   - Blue header (#1F4E78) matching export format")
print("   - 4 sample rows with realistic data")
print("   - All 10 columns: ID, Username, Email, First Name, Last Name, Middle Name, Role, Status, Password, Created Date")
print("   - Proper column widths and styling")
print("   - Frozen header row")
